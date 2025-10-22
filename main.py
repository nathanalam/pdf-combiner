import os
from pathlib import Path
from pypdf import PdfWriter, PdfReader
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
    KeepTogether,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfgen import canvas
from reportlab.platypus.tableofcontents import TableOfContents
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm


def convert_xlsx_to_pdf(xlsx_path: Path, output_path: Path) -> bool:
    """Convert an XLSX file to PDF format."""
    try:
        # Load the workbook
        wb = load_workbook(xlsx_path, data_only=True)

        # Create PDF with landscape orientation for better spreadsheet display
        # Use A4 landscape for more width
        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=landscape(A4),
            leftMargin=0.5 * inch,
            rightMargin=0.5 * inch,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
        )
        elements = []
        styles = getSampleStyleSheet()

        # Calculate available page width for table scaling
        available_width = landscape(A4)[0] - (1.0 * inch)  # Subtract margins

        # Create custom paragraph styles for table cells
        cell_style_normal = ParagraphStyle(
            "CellNormal",
            parent=styles["Normal"],
            fontSize=8,
            leading=10,
            alignment=TA_LEFT,
            spaceBefore=2,
            spaceAfter=2,
        )

        cell_style_header = ParagraphStyle(
            "CellHeader",
            parent=styles["Normal"],
            fontSize=9,
            leading=11,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
            spaceBefore=2,
            spaceAfter=2,
        )

        # Process each sheet
        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]

            # Add sheet name as header (skip if only one sheet)
            if len(wb.sheetnames) > 1:
                if sheet_idx > 0:
                    elements.append(PageBreak())
                header_style = styles["Heading1"]
                header_style.alignment = TA_CENTER
                elements.append(Paragraph(f"<b>Sheet: {sheet_name}</b>", header_style))
                elements.append(Spacer(1, 0.2 * inch))

            # Get all data from the sheet
            data = []
            max_col_count = 0

            for row in ws.iter_rows(values_only=True):
                # Convert None to empty string and all values to strings
                row_data = [str(cell) if cell is not None else "" for cell in row]
                # Skip completely empty rows
                if any(cell.strip() for cell in row_data):
                    data.append(row_data)
                    max_col_count = max(max_col_count, len(row_data))

            if data:
                # Ensure all rows have the same number of columns
                for row in data:
                    while len(row) < max_col_count:
                        row.append("")

                # Convert data to Paragraph objects for proper text wrapping
                paragraph_data = []
                for row_idx, row in enumerate(data):
                    paragraph_row = []
                    for col_idx, cell_content in enumerate(row):
                        # Clean up the content and handle special characters
                        content = str(cell_content).strip()
                        if not content:
                            content = " "  # Use space to maintain cell structure

                        # Choose appropriate style
                        if row_idx == 0:  # Header row
                            para = Paragraph(content, cell_style_header)
                        else:
                            para = Paragraph(content, cell_style_normal)

                        paragraph_row.append(para)
                    paragraph_data.append(paragraph_row)

                # Calculate column widths based on content
                col_widths = []
                for col_idx in range(max_col_count):
                    max_len = 0
                    for row in data:
                        cell_content = str(row[col_idx])
                        # Estimate width based on character count
                        max_len = max(max_len, len(cell_content))
                    # Minimum width, scale based on content but cap at reasonable size
                    col_width = min(max(1.2, max_len * 0.08), 3.5) * inch
                    col_widths.append(col_width)

                # Scale table to fit page width if needed
                total_width = sum(col_widths)
                if total_width > available_width:
                    scale_factor = available_width / total_width
                    col_widths = [w * scale_factor for w in col_widths]

                # Create table with calculated column widths and paragraph data
                table = Table(paragraph_data, colWidths=col_widths)

                # Enhanced styling for better readability
                table.setStyle(
                    TableStyle(
                        [
                            # Header row styling
                            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, 0), 9),
                            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                            ("TOPPADDING", (0, 0), (-1, 0), 8),
                            # Data row styling
                            ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                            ("FONTSIZE", (0, 1), (-1, -1), 8),
                            ("TOPPADDING", (0, 1), (-1, -1), 6),
                            ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
                            # Grid styling
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                            ("VALIGN", (0, 0), (-1, -1), "TOP"),
                            # Alternating row colors for better readability
                            (
                                "ROWBACKGROUNDS",
                                (0, 1),
                                (-1, -1),
                                [colors.white, colors.lightgrey],
                            ),
                        ]
                    )
                )

                # Wrap the table in KeepTogether to prevent page breaks within it
                elements.append(KeepTogether([table]))
                elements.append(Spacer(1, 0.3 * inch))

        # Build PDF
        doc.build(elements)
        return True
    except Exception as e:
        print(f"  Error converting {xlsx_path.name}: {e}")
        return False


def main():
    current_dir = Path(__file__).parent

    # Convert XLSX files to PDF first
    xlsx_files = sorted(current_dir.glob("*.xlsx"))
    converted_pdfs = []

    if xlsx_files:
        print(f"Found {len(xlsx_files)} XLSX file(s) to convert:")
        for xlsx_file in xlsx_files:
            print(f"  - {xlsx_file.name}")

        print(
            "\nConverting XLSX files to PDF with landscape orientation and text wrapping..."
        )
        for xlsx_file in xlsx_files:
            output_pdf = current_dir / f"{xlsx_file.stem}_converted.pdf"
            if convert_xlsx_to_pdf(xlsx_file, output_pdf):
                print(f"  Converted: {xlsx_file.name} -> {output_pdf.name}")
                converted_pdfs.append(output_pdf)
        print()

    # Get all PDF files (excluding the output file if it exists)
    all_pdf_files = sorted(current_dir.glob("*.pdf"))
    pdf_files = [pdf for pdf in all_pdf_files if pdf.name != "combined_output.pdf"]

    if not pdf_files:
        print("No PDF files found to merge.")
        return

    print(f"Found {len(pdf_files)} PDF file(s) to merge:")
    for pdf in pdf_files:
        marker = "(converted)" if pdf in converted_pdfs else ""
        print(f"  - {pdf.name} {marker}")

    # Create a PdfWriter object
    merger = PdfWriter()

    # Append each PDF
    print("\nMerging PDFs...")
    for pdf_file in pdf_files:
        try:
            reader = PdfReader(str(pdf_file))
            for page in reader.pages:
                merger.add_page(page)
            print(f"  Added: {pdf_file.name}")
        except Exception as e:
            print(f"  Error adding {pdf_file.name}: {e}")

    # Write the merged PDF
    output_file = current_dir / "combined_output.pdf"
    with open(output_file, "wb") as f:
        merger.write(f)

    print(f"\nSuccessfully combined {len(pdf_files)} PDFs into: {output_file.name}")


if __name__ == "__main__":
    main()
